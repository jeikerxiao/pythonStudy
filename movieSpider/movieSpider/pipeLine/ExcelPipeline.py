import openpyxl


class ExcelPipeline(object):

    def __init__(self):
        self.file_name = 'movieSpider/download/movies.xlsx'
        self.cur_row = 2
        self.creatwb(wbname=self.file_name)
        # 写入头信息
        title = ['rank', 'name', 'alias', 'rating_num', 'quote', 'url']
        sheet_name = 'movies'
        self.write_header(headers=title, sheetname=sheet_name, wbname=self.file_name)

    def process_item(self, item, spider):
        # 写入数据
        self.savetoexcel(item, wbname=self.file_name)
        return item

    # 新建excel
    def creatwb(self, wbname):
        wb = openpyxl.Workbook()
        wb.save(filename=wbname)

    # 写入表头信息
    def write_header(self, headers, sheetname, wbname):
        wb = openpyxl.load_workbook(filename=wbname)
        sheet = wb.active
        sheet.title = sheetname
        field = 1
        for field in range(1, len(headers) + 1):
            _ = sheet.cell(row=1, column=field, value=str(headers[field - 1]))
        wb.save(filename=wbname)

    # 写入excel文件中 item 数据
    def savetoexcel(self, item, wbname):
        wb = openpyxl.load_workbook(filename=wbname)
        sheet = wb.active
        sheet.cell(row=self.cur_row, column=1, value=str(item['rank']))
        sheet.cell(row=self.cur_row, column=2, value=str(item['name']))
        sheet.cell(row=self.cur_row, column=3, value=str(item['alias']))
        sheet.cell(row=self.cur_row, column=4, value=str(item['rating_num']))
        sheet.cell(row=self.cur_row, column=5, value=str(item['quote']))
        sheet.cell(row=self.cur_row, column=6, value=str(item['url']))
        self.cur_row += 1
        wb.save(filename=wbname)

